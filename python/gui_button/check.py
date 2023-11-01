import tkinter as tk
from tkinter import *
from tkinter import ttk,messagebox

from firebase import firebase
def check():
    contacts = []
    def close():
        fdb.put('/','ppt',8)
        query_wnd.destroy()
    def download():
        import os
        import openpyxl
        # 如果沒有對應excel檔就新增一個空白excel檔
        if os.path.exists('check.xlsx') == False:
            wb = openpyxl.Workbook()    # 建立空白的 Excel 活頁簿物件
            wb.save('check.xlsx')       # 儲存檔案
        # 新增資料
        wb = openpyxl.load_workbook('check.xlsx', data_only=True)

        wb.remove(wb['Sheet'])
        s3 = wb.create_sheet('Sheet')
        s3.append(["日期","事件", "收入",'支出'])
        
        datas = [list(a) for a in contacts]   # 二維陣列資料
        for i in datas:
            s3.append(i)                   # 逐筆添加到最後一列

        wb.save('check.xlsx')
    query_wnd  =  tk. Tk()
    query_wnd. geometry("550x300")
    query_wnd. title("我是收支表")
    url = "https://python-database-3b3f8-default-rtdb.firebaseio.com/"
    fdb = firebase.FirebaseApplication(url, None)
    fdb.put('/','ppt',15)
    # 实例化控件，设置表头样式和标题文本
    columns = ("日期","事件", "收入",'支出')
    headers = ("日期","事件", "收入",'支出')
    widthes = (125,135,125,125)
    tv = ttk.Treeview(query_wnd, show="headings", columns=columns)

    for (column, header, width) in zip(columns, headers, widthes):
        tv.column(column, width=width, anchor="w")
        tv.heading(column, text=header, anchor="w")
    
    def inser_data():
        """插入数据"""
        result = fdb.get('/money/event',None)
        
        for time,event in result.items():
            if event['type']=='outcome':
                contacts.append((time,event['event'],0,event['value']))
            else:
                contacts.append((time,event['event'],event['value'],0))
        for i, v in enumerate(contacts):
            print(i,v)
            tv.insert('', i,values=v)
    tv.pack()
    inser_data()
    result=fdb.get('/money',None)
    label_q = tk.Label(query_wnd,text='戶頭餘額:'+str(result['total']))
    label_q.place(relx=0.5,rely=0.8,anchor="center")
    download_button = tk.Button(query_wnd, text="下載報表",underline=-1,command=download)
    download_button.place(relx=0.3,rely=0.9,anchor="center")
    close_button = tk.Button(query_wnd, text="close",underline=-1,command=close)
    close_button.place(relx=0.7,rely=0.9,anchor="center")
    query_wnd.mainloop()